"""
Генерирует чистый xlsx-шаблон для ведения ставок.

Содержит:
- Лист 'Ставки' — основная таблица с правильными колонками и dropdown-валидацией
- Лист 'Справочники' — списки допустимых значений (можно править)
- Лист 'README' — инструкция как вести
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter


# ============================================================
# СПРАВОЧНИКИ (закрытые списки для дропдаунов)
# ============================================================
CHANNELS = [
    "Ночная Америка",
    "Канал для своих 2.0",
    "Платные экспрессы",
    "Сигнал / Лига ставок",
    "Сигналы «Первый забьёт»",
    "Бесплатные прогнозы",
]

SPORTS = [
    "Футбол",
    "Хоккей",
    "Баскетбол",
    "Бейсбол",
    "Волейбол",
    "Теннис",
    "MMA",
    "Киберспорт",
]

LEAGUES = [
    "NHL", "KHL", "AHL", "SHL",
    "NBA", "WNBA", "Euroleague",
    "MLB",
    "Premier League", "LaLiga", "Serie A", "Bundesliga", "Ligue 1",
    "РПЛ", "Кубок России",
    "MLS", "J1 League", "Saudi Pro League", "China Super League",
    "Portugal Primeira Liga", "Czech First League", "Bulgarian First League",
    "Kazakhstan Premier League", "Belgium Pro League",
    "Russia Volleyball Superleague",
    "Другое / разовая",
]

BET_TYPES = [
    # Хоккей / баскетбол / футбол — броски и подобное
    "Броски в створ, тотал",
    "Броски в створ, фора",
    "Броски в створ, победа",
    # Голы
    "Голы команды, индивидуальный тотал",
    "Голы команды, тотал",
    "Голы команды, фора",
    "Тотал матча, тотал",
    "Тотал матча, комбо",
    # Карты / фолы / угловые
    "Жёлтые карты, тотал",
    "Жёлтые карты, индивидуальный тотал",
    "Фолы, тотал",
    "Угловые, тотал",
    "Угловые, индивидуальный тотал",
    "Угловые, фора",
    # Баскетбол спецы
    "Очки команды, индивидуальный тотал",
    "Подборы, тотал",
    "Подборы, фора",
    # Бейсбол
    "Раны / иннинг, тотал",
    "Раны / иннинг, фора",
    # Исходы и спецрынки
    "Исход матча",
    "Двойной шанс",
    "Обе забьют",
    "Первый гол",
    "Фора по основному счёту",
    # Сборные
    "Экспресс",
    "Другое",
]

DIRECTIONS = [
    "больше",
    "меньше",
    "плюс",
    "минус",
    "ноль",
    "П1",
    "П2",
    "Х",
    "1Х",
    "Х2",
    "12",
    "да",
    "нет",
    "комбинированное",
    "—",  # для случаев без направления
]

RESULTS = ["WIN", "LOSE", "RETURN", "PENDING"]


# ============================================================
# КОЛОНКИ ОСНОВНОЙ ТАБЛИЦЫ
# ============================================================
# (название, ширина, обязательная?, dropdown_list_or_None, формат)
COLUMNS = [
    ("дата",          12, True,  None,       "yyyy-mm-dd"),
    ("канал",         24, True,  CHANNELS,   None),
    ("спорт",         13, True,  SPORTS,     None),
    ("лига",          22, True,  LEAGUES,    None),
    ("матч",          35, True,  None,       None),
    ("тип_ставки",    35, True,  BET_TYPES,  None),
    ("направление",   16, True,  DIRECTIONS, None),
    ("линия",          8, False, None,       "0.0"),
    ("ставка",        10, True,  None,       "0"),
    ("кэф",            8, True,  None,       "0.00"),
    ("результат",     12, True,  RESULTS,    None),
    ("профит",        10, True,  None,       "0;[Red]-0;-"),  # формула, считаем сами
    # Опциональные:
    ("кэф_прогноза",  12, False, None,       "0.00"),
    ("комиссия",      10, False, None,       "0"),
    ("формат",        10, False, ["SINGLE", "EXPRESS"], None),
    ("заметки",       40, False, None,       None),
]

N_DATA_ROWS = 500  # сколько строк подготовить с валидацией


# ============================================================
# СТИЛИ
# ============================================================
HEADER_FONT = Font(name="Arial", bold=True, color="FFFFFF", size=11)
HEADER_FILL_REQUIRED = PatternFill("solid", fgColor="2E5C8A")   # тёмно-синий — обязательная
HEADER_FILL_OPTIONAL = PatternFill("solid", fgColor="6B6B6B")   # серый — опциональная
HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
THIN_BORDER = Border(
    left=Side(style="thin", color="CCCCCC"),
    right=Side(style="thin", color="CCCCCC"),
    top=Side(style="thin", color="CCCCCC"),
    bottom=Side(style="thin", color="CCCCCC"),
)


# ============================================================
# СБОРКА ФАЙЛА
# ============================================================
def build_template(output_path: str):
    wb = Workbook()

    # ----- Лист 1: Ставки -----
    ws = wb.active
    ws.title = "Ставки"

    # Заголовки
    for col_idx, (name, width, required, _dropdown, _fmt) in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=name)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL_REQUIRED if required else HEADER_FILL_OPTIONAL
        cell.alignment = HEADER_ALIGN
        cell.border = THIN_BORDER
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.row_dimensions[1].height = 32
    ws.freeze_panes = "A2"  # шапка не уезжает при скролле

    # Форматы чисел и формула для профита
    profit_col_idx = next(i for i, c in enumerate(COLUMNS, start=1) if c[0] == "профит")
    stake_col_letter = get_column_letter(next(i for i, c in enumerate(COLUMNS, start=1) if c[0] == "ставка"))
    odds_col_letter = get_column_letter(next(i for i, c in enumerate(COLUMNS, start=1) if c[0] == "кэф"))
    result_col_letter = get_column_letter(next(i for i, c in enumerate(COLUMNS, start=1) if c[0] == "результат"))
    profit_col_letter = get_column_letter(profit_col_idx)

    for row in range(2, N_DATA_ROWS + 2):
        # Форматы
        for col_idx, (name, _w, _req, _dd, fmt) in enumerate(COLUMNS, start=1):
            if fmt:
                ws.cell(row=row, column=col_idx).number_format = fmt
        # Формула профита:
        # WIN -> stake * (odds-1), LOSE -> -stake, RETURN/PENDING -> 0
        formula = (
            f'=IF({result_col_letter}{row}="WIN",{stake_col_letter}{row}*({odds_col_letter}{row}-1),'
            f'IF({result_col_letter}{row}="LOSE",-{stake_col_letter}{row},'
            f'IF({result_col_letter}{row}="RETURN",0,"")))'
        )
        ws.cell(row=row, column=profit_col_idx, value=formula)

    # ----- Лист 2: Справочники (создаём ДО dropdownов, чтобы можно было на него ссылаться) -----
    ws_lists = wb.create_sheet("Справочники")
    ws_lists["A1"] = "Здесь хранятся справочники. Можно добавлять новые значения — они появятся в дропдаунах."
    ws_lists["A1"].font = Font(italic=True, color="888888")

    # Маппинг: имя колонки в основной таблице -> (буква столбца в Справочники, список значений)
    lists_data = [
        ("канал",         "Каналы",       CHANNELS),
        ("спорт",         "Спорты",       SPORTS),
        ("лига",          "Лиги",         LEAGUES),
        ("тип_ставки",    "Типы ставок",  BET_TYPES),
        ("направление",   "Направления",  DIRECTIONS),
        ("результат",     "Результаты",   RESULTS),
        ("формат",        "Форматы",      ["SINGLE", "EXPRESS"]),
    ]

    # column_name -> 'Справочники!$A$4:$A$N'  (ссылка для data validation)
    ref_by_column = {}

    for idx, (column_name, title, values) in enumerate(lists_data):
        col_letter = get_column_letter(idx + 1)
        # Заголовок столбца справочника
        ws_lists.cell(row=3, column=idx + 1, value=title).font = Font(bold=True, color="2E5C8A")
        ws_lists.column_dimensions[col_letter].width = max(len(title), max(len(v) for v in values)) + 2
        # Сами значения, начиная с 4-й строки
        for i, v in enumerate(values, start=4):
            ws_lists.cell(row=i, column=idx + 1, value=v)
        # Сохраняем ссылку для последующего dropdown'а — с запасом строк для будущих добавлений
        last_row = 4 + len(values) - 1 + 50  # +50 пустых строк на расширение
        ref_by_column[column_name] = f"Справочники!${col_letter}$4:${col_letter}${last_row}"

    # Dropdown'ы (data validation) — ссылаемся на ячейки в листе 'Справочники'
    for col_idx, (name, _w, _req, dropdown, _fmt) in enumerate(COLUMNS, start=1):
        if dropdown:
            col_letter = get_column_letter(col_idx)
            if name in ref_by_column:
                # Ссылка на диапазон в листе Справочники (работает с любой длиной)
                formula1 = f"={ref_by_column[name]}"
            else:
                # Fallback: inline-список (только если короткий)
                inline = '"' + ",".join(dropdown) + '"'
                if len(inline) > 250:
                    continue  # пропускаем, чтобы не сломать файл
                formula1 = inline
            dv = DataValidation(type="list", formula1=formula1, allow_blank=True)
            dv.add(f"{col_letter}2:{col_letter}{N_DATA_ROWS+1}")
            dv.error = "Выбери значение из списка"
            dv.errorTitle = "Неверное значение"
            ws.add_data_validation(dv)

    # ----- Лист 3: README -----
    ws_readme = wb.create_sheet("README", 0)  # вставим первым
    ws_readme.column_dimensions["A"].width = 22
    ws_readme.column_dimensions["B"].width = 90

    readme = [
        ("Назначение", "Простая таблица для учёта ставок и анализа эффективности прогнозов через Streamlit-дашборд."),
        ("Главные вопросы", "1) Какие каналы прогнозов реально зарабатывают?  2) Какие типы ставок сливают?"),
        ("", ""),
        ("📋 КАК ВЕСТИ", ""),
        ("1.", "Заполняй строки на листе «Ставки» — по одной строке на одну ставку."),
        ("2.", "Категориальные поля (синие заголовки) выбирай из выпадающих списков."),
        ("3.", "«профит» считается автоматически по результату — руками не трогай."),
        ("4.", "Если нужно добавить новую лигу/тип/канал — добавь в лист «Справочники» в соответствующий столбец."),
        ("", ""),
        ("🔵 ОБЯЗАТЕЛЬНЫЕ КОЛОНКИ", ""),
        ("дата",         "Дата матча, формат YYYY-MM-DD."),
        ("канал",        "Откуда прогноз (Ночная Америка, Канал для своих и т.д.). Это главная ось аналитики."),
        ("спорт",        "Вид спорта."),
        ("лига",         "Лига/чемпионат. Если разовая лига — выбери «Другое / разовая»."),
        ("матч",         "Названия команд, формат «Команда1 — Команда2». Для аналитики не используется, нужно для проверки."),
        ("тип_ставки",   "Что прогнозируется. Главная ось аналитики типов. Пример: «Броски в створ, фора»."),
        ("направление",  "больше/меньше для тоталов, плюс/минус для фор, П1/П2/Х для исходов."),
        ("линия",        "Численная линия (5.5, -1.5, +2.5). Для исходов и «обе забьют» — оставь пусто."),
        ("ставка",       "Размер ставки в рублях."),
        ("кэф",          "Коэффициент, по которому реально поставлена."),
        ("результат",    "WIN / LOSE / RETURN / PENDING. PENDING = ещё не сыграло."),
        ("профит",       "Автоматически считается: WIN → ставка×(кэф−1); LOSE → −ставка; RETURN/PENDING → 0."),
        ("", ""),
        ("⚪ ОПЦИОНАЛЬНЫЕ", ""),
        ("кэф_прогноза", "Кэф, который был в исходном прогнозе. Позволяет видеть, насколько ставка отличается от рекомендации."),
        ("комиссия",     "Если канал берёт % от выигрыша — сюда. Считается в ROI с учётом комиссии."),
        ("формат",       "SINGLE или EXPRESS. Если все ставки одиночные — можно не заполнять."),
        ("заметки",      "Свободный текст: контекст, ссылка на источник, причины расхождения."),
        ("", ""),
        ("📊 ЧТО ПОКАЖЕТ ДАШБОРД", ""),
        ("", "• Общий ROI и винрейт за выбранный период"),
        ("", "• ROI по каждому каналу прогноза (главное)"),
        ("", "• ROI по типу ставки"),
        ("", "• Тепловая карта канал × тип ставки — где жжёт, где сливает"),
        ("", "• Топ-10 прибыльных и убыточных подтипов"),
        ("", "• Подсветка «мало данных» (n < 30) — чтобы не делать выводы из шума"),
    ]

    ws_readme["A1"] = "ШАБЛОН УЧЁТА СТАВОК"
    ws_readme["A1"].font = Font(name="Arial", bold=True, size=16, color="2E5C8A")
    ws_readme.merge_cells("A1:B1")

    for i, (key, val) in enumerate(readme, start=3):
        cell_a = ws_readme.cell(row=i, column=1, value=key)
        cell_b = ws_readme.cell(row=i, column=2, value=val)
        if key and not val:  # заголовок секции
            cell_a.font = Font(bold=True, size=12, color="2E5C8A")
            ws_readme.merge_cells(start_row=i, start_column=1, end_row=i, end_column=2)
        elif key in ("Назначение", "Главные вопросы"):
            cell_a.font = Font(bold=True, color="2E5C8A")
            cell_b.alignment = Alignment(wrap_text=True, vertical="top")
        else:
            cell_a.font = Font(bold=True, color="555555")
            cell_b.alignment = Alignment(wrap_text=True, vertical="top")

    wb.save(output_path)
    print(f"Сохранён шаблон: {output_path}")


if __name__ == "__main__":
    build_template("/home/claude/stavki_app/template_stavki.xlsx")
